
import urllib.request as req
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Color, Side
import os
import time
# import win32com.client

# excel = win32com.client.Dispatch("Excel.Application")
# search_code = input("검색어를 입력하세요 >>>>>")

# search_code = "(urine) AND (albumin) AND ((biosensor) OR (sensor))"
search_code = "(((body fluids) OR (urine)) AND ((hydrogen ion) OR (nitrite) OR (albumin) OR (bilirubin) OR (hemoglobin) OR (glucose) OR (ketone) OR (leukocyte esterase) OR (specific gravity)) AND ((biosensor) OR (sensor))) NOT ((review) OR (DNA) OR (rna) or (cancer) OR (color) OR (photo) OR (optical) OR (dopamine) OR (colorimetric) OR (plasmon) OR (fluorescent) OR (electrochemiluminescence) OR (pseudomonas) OR (bladder))"
pubmed_code = search_code.replace(" ","+").replace("(","%28").replace(")","%29")
# search_year = input("언제부터 검색? >>")
start_year = 2016
end_year = 2021

pubmed_page_num = 1
headers = {'User-Agent': 'Chrome/66.0.3359.181'}
pub_url = "https://pubmed.ncbi.nlm.nih.gov/?term={}&filter=simsearch3.fft&filter=lang.english&filter=years.{}-{}&page={}".format(pubmed_code, start_year, end_year, pubmed_page_num)
pub_reqUrl = req.Request(pub_url, headers=headers)
code = req.urlopen(pub_reqUrl)
soup = BeautifulSoup(code, "lxml")
Paper_number = soup.select_one("div.results-amount span.value")
print()
print("Pubmed database에서")
print()
print("{} 키워드로 검색 결과".format(search_code))
print()
print("{}년부터 {}년까지 검색된 논문은 {}개 입니다.".format(start_year, end_year, Paper_number.text))
print()
deci_cro = int(input("논문 검색을 진행하시겠습니까? 진행:1  / 멈춤:2 >>>> "))
if deci_cro == 2:
    print("검색을 종료합니다")
else:
#엑셀 파일 만들기
    if not os.path.exists("./00_SearchPaper_result.xlsx"):
        book = openpyxl.Workbook()
        book.save("./00_SearchPaper_result.xlsx")
    #엑셀 파일 초기 설정
    book = openpyxl.load_workbook("./00_SearchPaper_result.xlsx")
    #pubmed 시트가 있으면 지우고 다시 만들고, 없으면 만들어서

    if not book["search_pubmed"]:
        book.create_sheet("search_pubmed")
        book.create_sheet("compare_result_pubmed")
    else:
        del book["search_pubmed"]
        book.create_sheet("search_pubmed")
        del book["compare_result_pubmed"]
        book.create_sheet("compare_result_pubmed")
    sheet_paper = book["Paper"]
    sheet_pub = book["search_pubmed"]
    sheet_pub.column_dimensions["A"].width = 15
    sheet_pub.column_dimensions["B"].width = 200
    sheet_compare = book["compare_result_pubmed"]
    sheet_compare.column_dimensions["A"].width = 30
    sheet_compare.column_dimensions["B"].width = 200
    myPaperData = 1
    while myPaperData < 24:
        sheet_pub.cell(row=myPaperData, column=1).value = sheet_paper.cell(row=myPaperData, column=1).value
        sheet_pub.cell(row=myPaperData, column=2).value = sheet_paper.cell(row=myPaperData, column=2).value
        myPaperData += 1
    #서식만들기
    user_font = Font(name="Times New Roman", size=12)
    yellow_color = PatternFill(patternType="solid", fgColor=Color("FFFF00"))
    user_alignment = Alignment(horizontal="center")
    sheet_pub["B19"].font =user_font
    sheet_pub["B23"].fill =yellow_color
    sheet_pub["B24"].fill =yellow_color
    sheet_pub["B25"].fill =yellow_color

    for row in sheet_pub["A1":"A1200"]:
        for cell in row:
            cell.font = user_font
            cell.alignment = user_alignment

    for row in sheet_pub["B1":"B1200"]:
        for cell in row:
            cell.font = user_font
    book.save("./00_SearchPaper_result.xlsx")
    #변수 정의
    # myPaperTitle = ["Recent Advances in Electric-Double-Layer Transistors for Bio-Chemical Sensing Applications","First Decade of Interfacial Iontronic Sensing: From Droplet Sensors to Artificial Skins"]
    # myPaperTitle = ["Quantitative determination of leukocyte esterase with a paper-based device","Electrical detection of blood cells in urine", "Electrochemical detection of nitrite ions using Ag/Cu/MWNT nanoclusters electrodeposited on a glassy carbon electrode", "Highly sensitive nitrite sensor based on AuNPs/RGO nanocomposites modified graphene electrochemical transistors", "IGZO-based electrolyte-gated field-effect transistor for in situ biological sensing platform", ""]
    #내 논문 가져오기
    myPaperTitle = []
    cell_range = sheet_paper["B1:B18"]
    for row in cell_range:
        for cell in row:
            myPaperTitle.append(cell.value)
    compare_result = []
    Paper_title_pubmed = []
     #내 논문빼고 20부터
    row_num = 26
    row_comp = 1
    row_excel = 1
    # while True:
    headers = {'User-Agent': 'Chrome/66.0.3359.181'}
    #==================================================================================================================================================
    ##펍메드 코드
    while True:
        pub_url = "https://pubmed.ncbi.nlm.nih.gov/?term={}&filter=simsearch3.fft&filter=lang.english&filter=years.{}-{}&page={}".format(pubmed_code, start_year,end_year, pubmed_page_num)
        pub_reqUrl = req.Request(pub_url, headers=headers)
        code = req.urlopen(pub_reqUrl)
        soup = BeautifulSoup(code, "lxml")
        Paper_title = soup.select("a.docsum-title")
        for i in Paper_title:
            text_striped = i.text.strip()
            text_striped_replaced = text_striped.replace(".","")
            # print(text_striped_replaced)
            sheet_pub.cell(row=row_num, column=2).value = text_striped_replaced
            sheet_pub.cell(row=row_num, column=1).value = row_excel
            row_num += 1
            row_excel += 1
            Paper_title_pubmed.append(text_striped_replaced)
            # print(row_num)
            #논문비교하기
        for title_i in Paper_title_pubmed:
            if title_i in myPaperTitle:
                compare_result.append(title_i)
        Paper_title_pubmed.clear()
        print("Pubmed 페이지 번호 = ", pubmed_page_num)
        pubmed_page_num += 1
        #더이상 논문이 없으면 탈출
        if len(Paper_title) == 0:
            break
        if pubmed_page_num > 99:
            print()
            print("1000개가 넘었습니다")
            break
    print("Pubmed에서 사용한 검색어는 {} 입니다".format(search_code))
    sheet_pub.cell(row=23, column=2).value = "Pubmed에서 찾은 논문 검색어는 {} 입니다".format(search_code)
    print("Pubmed에서 찾은 {}년부터 {}년까지 찾은 논문 개수는 {}개입니다.".format(start_year,end_year,row_excel-1))

    # #===============================================
    # #논문 비교
    for title_comp in compare_result:
        sheet_compare.cell(row=row_comp, column=1).value = "pubmed"
        sheet_compare.cell(row=row_comp, column=2).value = title_comp
        row_comp += 1
    # print("내 논문은 {}입니다".format(myPaperTitle))
    print("Pubmed에 포함된 내 논문은 {}개 입니다".format(row_comp-1))
    print("Pubmed에 포함된 내 논문은 {}입니다".format(compare_result))
    sheet_pub.cell(row=24, column=2).value = "Pubmed에서 찾은 {}년부터 {}년까지 찾은 전체 논문은 {}개, 그 중 내 논문은 {}개입니다".format(start_year,end_year,row_excel-1,row_comp-1)
    sheet_pub.cell(row=25, column=2).value = "Pubmed에 포함된 내 논문은 {}입니다".format(compare_result)
    # print(compare_result)
    #반드시 세이브 해야해
    book.save("./00_SearchPaper_result.xlsx")
    print()

    print("Pubmed 논문 검색 결과 작성이 완료되었습니다.")
    # excel.Visible = True
    # excel.Workbooks.Open("C:/Users/User/PycharmProjects/pythonProject1/00_SearchPaper_result.xlsx")
# ##펍메드 코드 완료
#==================================================================================================================================================
